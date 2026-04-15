import os
import json
import base64
import sqlite3
import httpx
import fitz  # PyMuPDF
from datetime import datetime
from io import BytesIO
from fastapi import FastAPI, File, UploadFile, Form, Request
from fastapi.responses import FileResponse, JSONResponse, Response
from fastapi.templating import Jinja2Templates
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = FastAPI(title="OCR Comptable")
templates = Jinja2Templates(directory="templates")

GLM_API_KEY = os.environ.get("GLM_API_KEY", "")
GLM_ENDPOINT = "https://open.bigmodel.cn/api/paas/v4/chat/completions"
DB_PATH = "/tmp/comptable.db"


# ─── DATABASE ────────────────────────────────────────────────────────────────

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS fatourat (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date_upload TEXT,
            fournisseur TEXT,
            date_facture TEXT,
            numero_facture TEXT,
            montant_ht TEXT,
            tva TEXT,
            total_ttc TEXT,
            notes TEXT,
            fichier_source TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS releves (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date_upload TEXT,
            compte TEXT,
            periode TEXT,
            date_op TEXT,
            libelle TEXT,
            debit TEXT,
            credit TEXT,
            solde TEXT,
            fichier_source TEXT
        )
    """)
    conn.commit()
    conn.close()

init_db()


# ─── HELPERS ─────────────────────────────────────────────────────────────────

def pdf_to_images(pdf_bytes: bytes) -> list:
    """Convert all PDF pages to list of base64 JPEG strings."""
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    images = []
    for page in doc:
        mat = fitz.Matrix(2.0, 2.0)  # 2× zoom = better OCR quality
        pix = page.get_pixmap(matrix=mat)
        img_bytes = pix.tobytes("jpeg")
        images.append(base64.b64encode(img_bytes).decode("utf-8"))
    doc.close()
    return images


def image_bytes_to_b64(data: bytes) -> str:
    return base64.b64encode(data).decode("utf-8")


def call_glm_vision(image_b64: str, doc_type: str) -> dict:
    if doc_type == "fatourah":
        prompt = (
            "Tu es un expert-comptable tunisien. Analyse cette facture et retourne "
            "UNIQUEMENT un objet JSON valide, sans markdown ni backticks:\n"
            '{"fournisseur":"","date_facture":"","numero_facture":"",'
            '"montant_ht":"","tva":"","total_ttc":"","notes":""}\n'
            "Montants: chiffres seuls, virgule décimale. Champ absent = chaîne vide."
        )
    else:
        prompt = (
            "Tu es un expert-comptable tunisien. Analyse ce relevé bancaire et retourne "
            "UNIQUEMENT un objet JSON valide, sans markdown ni backticks:\n"
            '{"compte":"","periode":"","operations":['
            '{"date":"","libelle":"","debit":"","credit":"","solde":""}]}\n'
            "Extrait TOUTES les lignes. Montants: chiffres seuls. Champ absent = chaîne vide."
        )

    headers = {
        "Authorization": f"Bearer {GLM_API_KEY}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": "glm-4v",
        "max_tokens": 3000,
        "messages": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:image/jpeg;base64,{image_b64}"},
                    },
                    {"type": "text", "text": prompt},
                ],
            }
        ],
    }

    resp = httpx.post(GLM_ENDPOINT, headers=headers, json=payload, timeout=90)
    resp.raise_for_status()
    text = resp.json()["choices"][0]["message"]["content"].strip()

    # Strip markdown fences if model ignores instructions
    if "```" in text:
        parts = text.split("```")
        for p in parts:
            p = p.strip()
            if p.startswith("json"):
                p = p[4:].strip()
            try:
                return json.loads(p)
            except Exception:
                continue

    return json.loads(text)


def merge_releve_results(pages: list) -> dict:
    """Merge multi-page relevé JSON objects into one."""
    merged = {"compte": "", "periode": "", "operations": []}
    for r in pages:
        if not merged["compte"] and r.get("compte"):
            merged["compte"] = r["compte"]
        if not merged["periode"] and r.get("periode"):
            merged["periode"] = r["periode"]
        ops = r.get("operations", [])
        if ops:
            merged["operations"].extend(ops)
    return merged


# ─── EXCEL GENERATORS ────────────────────────────────────────────────────────

THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def style_header(ws, col: int, value: str, fill_hex: str):
    cell = ws.cell(row=1, column=col, value=value)
    cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN
    return cell


def auto_col_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 45)


def generate_fatourat_excel() -> bytes:
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        "SELECT date_upload,fournisseur,date_facture,numero_facture,"
        "montant_ht,tva,total_ttc,notes,fichier_source FROM fatourat ORDER BY id DESC"
    )
    rows = c.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Factures"
    ws.row_dimensions[1].height = 30

    headers = [
        "Date Upload", "Fournisseur", "Date Facture", "N° Facture",
        "Montant HT", "TVA", "Total TTC", "Notes", "Fichier Source",
    ]
    for i, h in enumerate(headers, 1):
        style_header(ws, i, h, "1B4F72")

    even_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
    odd_fill = PatternFill(start_color="FDFEFE", end_color="FDFEFE", fill_type="solid")

    for ri, row in enumerate(rows, 2):
        fill = even_fill if ri % 2 == 0 else odd_fill
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.border = THIN
            cell.alignment = Alignment(vertical="center")

    auto_col_width(ws)
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_releves_excel() -> bytes:
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        "SELECT date_upload,compte,periode,date_op,libelle,"
        "debit,credit,solde,fichier_source FROM releves ORDER BY id DESC"
    )
    rows = c.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relevés Bancaires"
    ws.row_dimensions[1].height = 30

    headers = [
        "Date Upload", "Compte", "Période", "Date Opération",
        "Libellé", "Débit", "Crédit", "Solde", "Fichier Source",
    ]
    for i, h in enumerate(headers, 1):
        style_header(ws, i, h, "154360")

    even_fill = PatternFill(start_color="EAF2FF", end_color="EAF2FF", fill_type="solid")
    odd_fill = PatternFill(start_color="FDFEFE", end_color="FDFEFE", fill_type="solid")

    for ri, row in enumerate(rows, 2):
        fill = even_fill if ri % 2 == 0 else odd_fill
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.border = THIN
            cell.alignment = Alignment(vertical="center")

    auto_col_width(ws)
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── ROUTES ──────────────────────────────────────────────────────────────────

@app.get("/", response_class=Response)
async def index(request: Request):
    with open("templates/index.html", "r", encoding="utf-8") as f:
        html = f.read()
    return Response(content=html, media_type="text/html")


@app.get("/stats")
async def stats():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM fatourat")
    nb_fat = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM releves")
    nb_rel = c.fetchone()[0]
    conn.close()
    return {"fatourat": nb_fat, "releves": nb_rel}


@app.post("/upload")
async def upload(
    file: UploadFile = File(...),
    doc_type: str = Form(...),  # "fatourah" or "releve"
):
    if not GLM_API_KEY:
        return JSONResponse({"error": "GLM_API_KEY non configurée sur le serveur."}, status_code=500)

    try:
        raw = await file.read()
        filename = file.filename or "document"
        now = datetime.now().strftime("%Y-%m-%d %H:%M")

        # ── Determine images to process
        if filename.lower().endswith(".pdf"):
            images = pdf_to_images(raw)
        else:
            images = [image_bytes_to_b64(raw)]

        # ── Call GLM for each page/image
        page_results = []
        for img_b64 in images:
            result = call_glm_vision(img_b64, doc_type)
            page_results.append(result)

        # ── Persist + build response
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()

        if doc_type == "fatourah":
            # For multi-page fatourahs, take first non-empty result
            data = page_results[0] if page_results else {}
            for r in page_results:
                if r.get("total_ttc") or r.get("fournisseur"):
                    data = r
                    break

            c.execute(
                "INSERT INTO fatourat (date_upload,fournisseur,date_facture,numero_facture,"
                "montant_ht,tva,total_ttc,notes,fichier_source) VALUES (?,?,?,?,?,?,?,?,?)",
                (
                    now,
                    data.get("fournisseur", ""),
                    data.get("date_facture", ""),
                    data.get("numero_facture", ""),
                    data.get("montant_ht", ""),
                    data.get("tva", ""),
                    data.get("total_ttc", ""),
                    data.get("notes", ""),
                    filename,
                ),
            )
            conn.commit()
            conn.close()
            return {"type": "fatourah", "data": data, "pages": len(images)}

        else:  # releve
            merged = merge_releve_results(page_results)
            ops = merged.get("operations", [])
            for op in ops:
                c.execute(
                    "INSERT INTO releves (date_upload,compte,periode,date_op,libelle,"
                    "debit,credit,solde,fichier_source) VALUES (?,?,?,?,?,?,?,?,?)",
                    (
                        now,
                        merged.get("compte", ""),
                        merged.get("periode", ""),
                        op.get("date", ""),
                        op.get("libelle", ""),
                        op.get("debit", ""),
                        op.get("credit", ""),
                        op.get("solde", ""),
                        filename,
                    ),
                )
            conn.commit()
            conn.close()
            return {
                "type": "releve",
                "data": merged,
                "rows_inserted": len(ops),
                "pages": len(images),
            }

    except json.JSONDecodeError as e:
        return JSONResponse({"error": f"GLM a retourné un JSON invalide: {str(e)}"}, status_code=422)
    except httpx.HTTPStatusError as e:
        return JSONResponse({"error": f"Erreur API GLM: {e.response.status_code} — {e.response.text}"}, status_code=502)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/export/fatourat")
async def export_fatourat():
    data = generate_fatourat_excel()
    fname = f"fatourat_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/export/releves")
async def export_releves():
    data = generate_releves_excel()
    fname = f"releves_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.delete("/clear/{table}")
async def clear_table(table: str):
    if table not in ("fatourat", "releves"):
        return JSONResponse({"error": "Table inconnue"}, status_code=400)
    conn = sqlite3.connect(DB_PATH)
    conn.execute(f"DELETE FROM {table}")
    conn.commit()
    conn.close()
    return {"ok": True, "table": table}
